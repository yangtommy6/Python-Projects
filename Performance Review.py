import requests
import openpyxl 
import mysql.connector
import csv

#print('Beginning file download with requests')

#download P1 Excel sheet from Box
url = 'https://byu.box.com/shared/static/kj42c56krg39j8i8xkgfy6pogz5rijk1.xlsx'
r = requests.get(url)



#save excel sheet to the server
with open('/home/ops/performanceReviews/box/christianP1.xlsx', 'wb') as f:
    f.write(r.content)

# Retrieve HTTP meta-data
print(r.status_code)
#print(r.headers['content-type'])
#print(r.encoding)




  
# Give the location of the file 
path = "/home/ops/performanceReviews/box/christianP1.xlsx"
 
  # workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row
  
# Loop will print all columns name 
for i in range(1, max_col + 1): 
    for y in range(1, max_row + 1):
        cell_obj = sheet_obj.cell(row = y, column = i) 
        # print(cell_obj.value) 
        

import pandas as pd
import os
df = pd.read_excel(
     os.path.join(path),
     engine='openpyxl',
     #locate to sheet 1--->
     sheet_name='Sheet1'
)
df = df[['Number','Date','Service/Problem','MIC','Contact Attempt','Contact Reasoning','OPS Resolution','Good Comms','Twitter/notify','Root Cause','Template','Timeline','Outage','Wrap Up <30','EvaluateTime (min)','Total','Notes']]

df.dropna(inplace = True)
print(df)
#df.to_csv (r'/home/ops/performanceReviews/box/christianP1.csv', index = None, header=True)
#df = pd.read_csv('/home/ops/performanceReviews/box/christianP1.csv')

mydb = mysql.connector.connect(
  host="localhost",
  user="byuopsne_wp1",
  password="X.Qxe04gHqA8wjYgdhB83",
  database="byuopsne_wp1"
)

try:

    if mydb.is_connected():
        cursor = mydb.cursor()
        cursor.execute("select database();")
        record = cursor.fetchone()
        #print("You're connected to database: ", record)
        

        for i,row in df.iterrows():
            sql = "INSERT IGNORE INTO christianP1 (number, date, service, mic, contactAttempt, contactReasoning, opsResolution, comms, twitterNotify, rootCause, template, timeline, outage, wrapUP, evaluateTime, total, notes, excuse) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'No')"
            cursor.execute(sql, tuple(row))
            print("Record inserted")
            # the connection is not autocommitted by default, so we must commit to save our changes
            mydb.commit()
except Exception as e:
    print("Error while connecting to MySQL", e)
